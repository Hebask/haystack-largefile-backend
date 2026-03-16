from openpyxl import load_workbook


def extract_xlsx_blocks(file_path: str, rows_per_block: int = 100):
    wb = load_workbook(file_path, read_only=True, data_only=True)
    blocks = []
    block_index = 1

    for sheet in wb.worksheets:
        current_rows = []
        current_count = 0

        for row in sheet.iter_rows(values_only=True):
            row_values = [str(v) if v is not None else "" for v in row]
            row_text = " | ".join(row_values).strip()

            if row_text:
                current_rows.append(row_text)
                current_count += 1

            if current_count >= rows_per_block:
                text = "\n".join(current_rows).strip()
                if text:
                    blocks.append(
                        {
                            "sheet_name": sheet.title,
                            "block_number": block_index,
                            "text": text,
                        }
                    )
                    block_index += 1

                current_rows = []
                current_count = 0

        if current_rows:
            text = "\n".join(current_rows).strip()
            if text:
                blocks.append(
                    {
                        "sheet_name": sheet.title,
                        "block_number": block_index,
                        "text": text,
                    }
                )
                block_index += 1

    wb.close()
    return blocks